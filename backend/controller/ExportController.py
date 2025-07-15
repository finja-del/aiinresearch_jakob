from fastapi import APIRouter, HTTPException
from backend.models.PaperDTO import PaperDTO
from typing import List
import os
import csv
from datetime import datetime
from fastapi import Request
from fastapi.responses import FileResponse
from fastapi.responses import StreamingResponse
from io import StringIO
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from fastapi import APIRouter, Request
from typing import List
from backend.models.PaperDTO import PaperDTO
from fastapi import UploadFile, File, Form
import json
import tempfile
import pandas as pd
from backend.models.PaperDTO import PaperDTO
from backend.models.FilterCriteria import FilterCriteriaIn
from backend.services.Deduplication.DuplicateMergeService import DuplicateMergeService
from backend.services.Filterservices.AbdcService import AbdcService
from backend.services.Filterservices.VhbService import VhbService

# Router definieren
router = APIRouter()

# Export-Verzeichnis definieren
EXPORT_DIR = os.path.join(os.path.dirname(__file__), '..', 'exports')
os.makedirs(EXPORT_DIR, exist_ok=True)



router = APIRouter()

@router.post("/download")
async def export_papers_as_xlsx(request: Request):
    data = await request.json()
    papers: List[PaperDTO] = [PaperDTO(**paper) for paper in data]

    # Neues Excel Workbook erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Papers"

    # Spaltenüberschriften
    headers = [
        'title', 'authors', 'abstract', 'date', 'source','sources','source_count',
        'vhbRanking', 'abdcRanking', 'journal_name', 'issn', 'eissn',
        'doi', 'url', 'citations', 'journal_quartile'
    ]

    # Überschriften in fett
    header_font = Font(bold=True)
    ws.append(headers)
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = header_font
            cell.alignment = cell.alignment.copy(wrap_text=True)


    # Inhalte einfügen
    for paper in papers:
        paper_dict = paper.to_api_dict()
        row = [
            paper_dict.get("title", ""),
            ", ".join(paper_dict.get("authors", [])) if isinstance(paper_dict.get("authors"), list) else paper_dict.get("authors", ""),
            paper_dict.get("abstract", ""),
            paper_dict.get("date", ""),
            paper_dict.get("source", ""),
            ", ".join(paper_dict.get("sources", [])) if isinstance(paper_dict.get("sources"), list) else paper_dict.get("sources", ""),
            paper_dict.get("source_count", 0),
            paper_dict.get("vhbRanking", ""),
            paper_dict.get("abdcRanking", ""),
            paper_dict.get("journal_name", ""),
            paper_dict.get("issn", ""),
            paper_dict.get("eissn", ""),
            paper_dict.get("doi", ""),
            paper_dict.get("url", ""),
            paper_dict.get("citations", 0),
            paper_dict.get("journal_quartile", "")
        ]
        ws.append(row)
    end_row = ws.max_row
    end_col = ws.max_column
    table_range = f"A1:{ws.cell(row=end_row, column=end_col).coordinate}"

    excel_table = Table(displayName="PaperExport", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    excel_table.tableStyleInfo = style
    ws.add_table(excel_table)
    # Spaltenbreiten automatisch anpassen (optional rudimentär)
    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    # Excel-Datei in BytesIO speichern
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=papers_export.xlsx"
        }
    )


@router.post("/uploadfile")
async def upload_file_and_process(
        file: UploadFile = File(...),
        filters: str = Form(...)
):
    try:
        # 👇 Debug-Marker – sollten immer erscheinen
        print("🟡 1. Endpoint HIT", flush=True)

        suffix = ".xlsx" if file.filename.endswith(".xlsx") else ".csv"
        print(f"🟡 2. suffix={suffix}", flush=True)

        # --- Datei speichern ---
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await file.read())
            tmp_path = tmp.name
        print(f"🟡 3. saved to {tmp_path}", flush=True)

        # Excel oder CSV laden
        if suffix == ".xlsx":
            df = pd.read_excel(tmp_path)
        else:
            # --- CSV einlesen mit automatischer Delimiter-Erkennung  ---
            with open(tmp_path, "r", encoding="utf-8-sig") as fh:
                sample = fh.read(4096)
            # Scopus: Komma-getrennt,  WoS: Semikolon oder Tab
            if sample.count(",") >= sample.count(";") and sample.count(",") >= sample.count("\t"):
                sep = ","
            elif sample.count(";") >= sample.count("\t"):
                sep = ";"
            else:
                sep = "\t"
            print(f"📑 CSV-Delimiter erkannt: '{sep}'")
            df = pd.read_csv(tmp_path, sep=sep, engine="python")

        # --- Spalten aus Scopus/WoS auf interne Feldnamen mappen ---
        column_aliases = {
            'Title': 'title',
            '"Title"': 'title',
            'Authors': 'authors',
            '"Authors"': 'authors',
            'Author full names': 'authors',
            '"Author full names"': 'authors',
            'Abstract': 'abstract',
            '"Abstract"': 'abstract',
            'Year': 'date',
            '"Year"': 'date',
            'Source title': 'journal_name',
            '"Source title"': 'journal_name',
            'DOI': 'doi',
            '"DOI"': 'doi',
            'Link': 'url',
            '"Link"': 'url',
            'Cited by': 'citations',
            '"Cited by"': 'citations'
        }
        df.rename(columns={c: column_aliases[c] for c in df.columns if c in column_aliases}, inplace=True)
        print("📦 Spalten nach Rename:", df.columns.tolist())

        # Falls nur Jahr vorhanden, mache daraus ein ISO-Datum
        if 'date' in df.columns:
            df['date'] = df['date'].apply(lambda y: f"{int(y)}-01-01" if pd.notna(y) else None)

        # Filter aus JSON extrahieren
        filter_obj = FilterCriteriaIn(**json.loads(filters))
        selected_ratings = set((filter_obj.rating or []) + (filter_obj.ranking or []))

        PAPERDTO_FIELDS = [
            'title', 'authors', 'abstract', 'date', 'source', 'sources', 'source_count',
            'vhbRanking', 'abdcRanking', 'journal_name', 'issn', 'eissn',
            'doi', 'url', 'citations', 'journal_quartile'
        ]
        # Rankings, Quellen, PaperDTOs bauen
        papers = []
        vhb_service = VhbService()
        abdc_service = AbdcService()

        for _, row in df.iterrows():
            paper_dict = row.to_dict()
            paper_dict = {k: v for k, v in paper_dict.items() if k in PAPERDTO_FIELDS}

            issn = paper_dict.get("issn") or ""
            journal_name = paper_dict.get("journal_name") or ""
            paper_dict["vhbRanking"] = vhb_service.getRanking(journal_title=journal_name, issn=issn)
            paper_dict["abdcRanking"] = abdc_service.getRanking(journal_title=journal_name, issn=issn)

            # Hauptquelle setzen: aus "source" oder "sources"
            if paper_dict.get("source") and paper_dict["source"] not in ["", "ManualUpload"]:
                pass  # bleibt wie es ist
            elif paper_dict.get("sources"):
                if isinstance(paper_dict["sources"], str):
                    paper_dict["sources"] = [s.strip() for s in paper_dict["sources"].split(",")]
                if isinstance(paper_dict["sources"], list) and paper_dict["sources"]:
                    paper_dict["source"] = paper_dict["sources"][0]
                else:
                    paper_dict["source"] = ""
            else:
                paper_dict["source"] = ""
                paper_dict["sources"] = []

            # sources immer als Liste!
            if isinstance(paper_dict.get("sources"), str):
                paper_dict["sources"] = [s.strip() for s in paper_dict["sources"].split(",")]
            elif not paper_dict.get("sources"):
                paper_dict["sources"] = [paper_dict["source"]] if paper_dict["source"] else []

            paper_dict["source_count"] = len(set(paper_dict["sources"]))

            papers.append(PaperDTO(**paper_dict))

        print("📄 Erzeugte PaperDTOs:", len(papers))

        # ======= MERGE-SERVICE anwenden =======
        try:
            merged_papers =  DuplicateMergeService.merge_duplicates(papers)
            print(f"📄 Nach Merge: {len(merged_papers)} Paper übrig.")
        except Exception as e:
            print("❌ MergeService Fehler:", e)
            merged_papers = papers  # fallback

        # --- Clean all float('nan') recursively to None for JSON serialization
        import math

        def clean_dict(obj):
            if isinstance(obj, dict):
                return {k: clean_dict(v) for k, v in obj.items()}
            elif isinstance(obj, float) and math.isnan(obj):
                return None
            elif isinstance(obj, list):
                return [clean_dict(i) for i in obj]
            return obj

        return clean_dict(merged_papers)

    except Exception as e:
        print(f"❌ Fehler beim Verarbeiten: {str(e)}")
        return {"error": f"Fehler beim Verarbeiten: {str(e)}"}


@router.post("/export")
async def export_papers_endpoint(request: Request):
     data = await request.json()
     controller = ExportController()
     papers: List[PaperDTO] = [PaperDTO(**paper) for paper in data]

     print("Erhaltene Daten:", data)
     filename = controller.export_papers_to_csv(papers)
     return {"status": "received"}

   #
   # return {"message": "Export erfolgreich.", "filename": filename}

class ExportController:

    def __init__(self):
        self.export_path = EXPORT_DIR

    def export_papers_to_csv(self, papers: List[PaperDTO]) -> str:
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"papers_export_{now}.csv"
        full_path = os.path.join(self.export_path, filename)

        # Header aus PaperDTO Feldern
        header = [
            'title', 'authors', 'abstract', 'date', 'source','sources','sourceCount', 'vhbRanking', 'abdcRanking',
            'journal_name', 'issn', 'eissn', 'doi', 'url', 'citations', 'journal_quartile'
        ]

        with open(full_path, mode="w", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=header, delimiter=';')
            writer.writeheader()

            for paper in papers:
                # Konvertiere PaperDTO in API-Dict für Export
                paper_dict = paper.to_api_dict()

                # Autoren als Komma-getrennte Zeichenkette
                if isinstance(paper_dict['authors'], list):
                    paper_dict['authors'] = ", ".join(paper_dict['authors'])

                writer.writerow(paper_dict)

        print(f"✅ Export abgeschlossen: {full_path}")
        return filename